from __future__ import annotations

from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .settings import COLUMNS


class ExcelWriter:
    _HFILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    _HFONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    _CFONT = Font(name="Calibri", size=10)
    _BORDER = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))

    def __init__(self, path: Path):
        self.path = path

    @staticmethod
    def _create_empty_workbook(path: Path) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        ws.title = "GEM Bids"
        ws.append(COLUMNS)
        for cell in ws[1]:
            cell.font = ExcelWriter._HFONT
            cell.fill = ExcelWriter._HFILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = ExcelWriter._BORDER
        wb.save(path)
        wb.close()

    @staticmethod
    def _current_header(ws) -> list[str]:
        return [str(c.value or "").strip() for c in ws[1][: len(COLUMNS)]]

    def _ensure_layout(self) -> None:
        if not self.path.exists():
            self._create_empty_workbook(self.path)
            return

        wb = load_workbook(self.path)
        ws = wb.active
        if self._current_header(ws) == COLUMNS:
            wb.close()
            return

        # Header mismatch: recreate with correct columns
        wb.close()
        self._create_empty_workbook(self.path)

    def _existing_refs(self) -> set[str]:
        self._ensure_layout()
        if not self.path.exists():
            return set()
        wb = load_workbook(self.path)
        ws = wb.active
        col = next((i for i, c in enumerate(ws[1], 1) if c.value == "Reference No."), None)
        if not col:
            wb.close()
            return set()
        refs = {
            str(r[0].value).strip()
            for r in ws.iter_rows(min_row=2, min_col=col, max_col=col)
            if r[0].value
        }
        wb.close()
        return refs

    def save(self, bids: list[dict]) -> int:
        self._ensure_layout()
        refs = self._existing_refs()
        new_bids = [b for b in bids if b.get("Reference No.") and b["Reference No."] not in refs]
        if not new_bids:
            if not self.path.exists():
                self._create_empty_workbook(self.path)
            return 0

        if self.path.exists():
            wb = load_workbook(self.path)
            ws = wb.active
        else:
            self._create_empty_workbook(self.path)
            wb = load_workbook(self.path)
            ws = wb.active

        for bid in new_bids:
            ws.append([bid.get(col, "") for col in COLUMNS])
            for c in ws[ws.max_row]:
                c.font = self._CFONT
                c.border = self._BORDER
                c.alignment = Alignment(vertical="center", wrap_text=True)

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        ws.auto_filter.ref = ws.dimensions
        wb.save(self.path)
        wb.close()
        return len(new_bids)
