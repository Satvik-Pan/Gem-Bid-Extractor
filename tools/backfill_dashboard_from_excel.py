from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from src.gem_bid_extractor.supabase_store import SupabaseStore

ROOT = Path(__file__).resolve().parents[1]
EXTRACTED_FILE = ROOT / "output" / "Extracted_bids.xlsx"
DOUBTFUL_FILE = ROOT / "output" / "doubtful_bids.xlsx"


def _rows_from_legacy_excel(path: Path, category: str) -> list[dict]:
    if not path.exists():
        return []

    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows: list[dict] = []

    for values in ws.iter_rows(min_row=2, values_only=True):
        vals = list(values)
        ref = str(vals[1] or "").strip() if len(vals) > 1 else ""
        if not ref:
            continue

        title = str(vals[0] or vals[2] or "").strip() if len(vals) > 2 else str(vals[0] or "").strip()
        desc = str(vals[3] or "").strip() if len(vals) > 3 else ""
        dept = str(vals[11] or vals[4] or "").strip() if len(vals) > 11 else (str(vals[4] or "").strip() if len(vals) > 4 else "")

        conf_raw = vals[13] if len(vals) > 13 else None
        try:
            conf = float(conf_raw) if conf_raw is not None else 0.7
        except (TypeError, ValueError):
            conf = 0.7

        rows.append(
            {
                "Reference No.": ref,
                "Bid ID": ref,
                "Name": title,
                "Description": desc,
                "Department": dept,
                "Pipeline Source": "legacy_excel_backfill",
                "LLM Confidence": conf,
                "LLM Reason": "Backfilled from local Excel output",
                "Final Category": category,
            }
        )

    wb.close()
    return rows


def main() -> None:
    extracted_rows = _rows_from_legacy_excel(EXTRACTED_FILE, "EXTRACTED")
    doubtful_rows = _rows_from_legacy_excel(DOUBTFUL_FILE, "DOUBTFUL")
    rows = [*extracted_rows, *doubtful_rows]

    store = SupabaseStore()
    ok = store.sync_with_retry(rows)
    print(f"Backfill rows: {len(rows)}")
    print(f"Supabase sync ok: {ok}")
    if store.last_error:
        print(f"Last error: {store.last_error}")


if __name__ == "__main__":
    main()
